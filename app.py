import pandas as pd
from datetime import datetime
from http import cookies
import json
from multiprocessing.dummy import Manager
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import random
import smtplib
import ssl
import flask
from flask import Flask
from flask import Flask, redirect,render_template,request, session, jsonify,url_for,send_from_directory,send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import DateTime, func,exc
from sqlalchemy.exc import IntegrityError
from werkzeug.utils import secure_filename
import openpyxl
import os


app = Flask(__name__)
app.secret_key='P@ju3125'

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///pixelstat.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

alert = None

uploads_dir = os.path.join(app.instance_path, 'uploads')
os.makedirs(uploads_dir, exist_ok=True)

class Admin(db.Model):
    username = db.Column(db.String(50), primary_key=True)
    password = db.Column(db.String(16), nullable=False)

class Managers(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False)
    email = db.Column(db.String(50),unique=True, nullable=False)
    password = db.Column(db.String(16), nullable=False)
    mobile = db.Column(db.Integer,unique=True, nullable=False)
    able_to_download = db.Column(db.Boolean, default=False, nullable=False)
    
    # def __repr__(self) -> str:
    #     return f"{self.id} - {self.name}"

class Region(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    region = db.Column(db.String(20), nullable=False)
    mid = db.Column(db.Integer, db.ForeignKey('managers.id'))

class LoginHistory(db.Model):
    id = db.Column(db.Integer, primary_key = True)
    mid = db.Column(db.Integer, db.ForeignKey('managers.id'), nullable = False)
    login_time = db.Column(db.DateTime, nullable = False)
    
    def __repr__(self):
        return f'<LoginHistory Manager ID: {self.mid} - Login Time: {self.login_time}>'
    
class Excel_data(db.Model):
    fno = db.Column(db.Integer, nullable=True)
    member_no  =db.Column(db.Integer, primary_key=True)
    membership_type = db.Column(db.String(100),default=0, nullable=True)
    company_name = db.Column(db.String(100),default=0, nullable=True)
    member_name = db.Column(db.String(100),default=0, nullable=True)
    contact = db.Column(db.Integer,default=0, nullable=True)
    email = db.Column(db.String(50),default=0, nullable=True)
    address = db.Column(db.String(200),default=0, nullable=True)
    dob = db.Column(DateTime, nullable=True)
    member_Category = db.Column(db.String(100),default=0, nullable=True)
    membership_date = db.Column(DateTime, nullable=True)
    region = db.Column(db.String(50),default=0, nullable=True)
    renewal_status = db.Column(db.Boolean, default=False, nullable=True)
    renewal_date = db.Column(DateTime, nullable=True)
    renewal_amount = db.Column(db.Integer,default=0, nullable=True)

with app.app_context():
    # db.drop_all()
    db.create_all()
    # a = Admin(username='admin', password='Admin@123')
    # db.session.add(a)
    db.session.commit()
    
# Redirect to login page
@app.route('/')
@app.route('/login')
def login():
    global alert
    alert=None
    return render_template('login.html')

@app.route('/verifyAdmin', methods=['GET','POST'])
def verifyUsername():
    global alert
    alert=None
    if request.method=='POST':
        login_data = request.get_json()
        username = login_data[0]['username']
        password = login_data[1]['password']
        try:
            user = Admin.query.filter_by(username=username,password=password).first()
            if user != None:
                login={
                    'user':'admin',
                    'username':username,
                    'status':True
                }
                alert = {
                    'type':'success',
                    'message':'Logged in successfully',
                    'section':'admin'
                }
                return jsonify(login)
            else:
                print(username)
                user = Managers.query.filter_by(mobile=int(username), password=password).first()
                print(user.id)
                if user != None:
                    login_time = datetime.now()
                    login_history = LoginHistory(mid = user.id, login_time = login_time) 
                    db.session.add(login_history)
                    db.session.commit()
                    # otp = sendOTP(username)
                    login={
                        'user':'manager',
                        'username':username,
                        # 'otp':otp,
                        'status':True
                    }
                    alert = {
                        'type':'success',
                        'message':'Logged in successfully',
                        'section':'manager'
                    }
                else:
                    login={
                        'status':False
                    }
                return jsonify(login)
        except:
            login={
                'status':False
            }
            return jsonify(login)
# To send OTP
def sendOTP(email):
    otp = random.randrange(100000,999999)
            
    sender_email = 'Your Email'
    password = 'Your Password'
    receiver_email = email

    message = MIMEMultipart("alternative")
    message["Subject"] = "PixelStat ERP : OTP (One Time Password)"
    message["From"] = sender_email
    message["To"] = receiver_email

    # Create the plain-text and HTML version of your message
    text = """\
    Verify your Email address to login ERP"""
    html = f"""\
    <html>
    <body>
        <h3>Verify your login using OTP</h3>
        <br>
        <h3>OTP : {otp}</h3>
    </body>
    </html>
    """

    part1 = MIMEText(text, "plain")
    part2 = MIMEText(html, "html")

    message.attach(part1)
    message.attach(part2)

    # Create secure connection with server and send email
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(sender_email, password)
        server.sendmail(
            sender_email, receiver_email, message.as_string()
        )
    return otp


def getData(page):
    allManagers = Managers.query.all()
    allregions = {}
    for manager in allManagers:
        regions = db.session.query(Region.region).filter_by(mid = manager.id).all()
        regions=list(regions)
        str=[]
        for region in regions:
            str.append(region[0])
        allregions[manager.id] = ', '.join(str)
    if 'filter' in request.cookies:
        filter = json.loads(request.cookies['filter'])                    
        if (filter['region'] == None and filter['membership_type']==None and filter['member_no']==None):
            if page == None:
                allRecords = Excel_data.query.all()
            else:
                allRecords = Excel_data.query.paginate(page=page,per_page=10)
        else:
            filtered = {k: v for k, v in filter.items() if v is not None}
            filter.clear()
            filter.update(filtered)
            filteredData = db.session.query(Excel_data)
            for attr, value in filter.items():
                filteredData = filteredData.filter(getattr(Excel_data, attr).like("%%%s%%" % value))
            
            if page == None:
                return filteredData, allManagers, allregions, alert
            else:
                return filteredData.paginate(page=page,per_page=10), allManagers, allregions, alert
    
    if page == None:
        allRecords = Excel_data.query.all()
    else:
        allRecords = Excel_data.query.paginate(page=page,per_page=10)
    return allRecords, allManagers, allregions, alert

# To delete all records
@app.route('/admin/deleteAllRecords', endpoint='deleteAllRecords')
def deleteRecords():
    if 'loggeduser' in request.cookies:
        loggeduser = json.loads(request.cookies['loggeduser'])
        if loggeduser['user'] == 'admin':
            try:
                pixelstat = Excel_data.query.delete()
                db.session.commit()
                alert={
                    'type':'success',
                    'message':'Records deleted successfully','section':'admin'
                }
            except:
                db.session.rollback()
                alert={
                    'type':'danger',
                    'message':'Something went wrong','section':'admin'
                }   
    return redirect('/admin#home')

# To create excel file
@app.route('/admin/downloadFile', endpoint='downloadFile')
@app.route('/manager/downloadFile',endpoint="downloadFile")
def createExcel():
    if 'loggeduser' in request.cookies:
        loggeduser = json.loads(request.cookies['loggeduser'])
        if loggeduser['user'] == 'admin':
            allRecords,allManagers,allregions,alert = getData(None)
        elif loggeduser['user'] == 'manager':
            allRecords, allregions,manager = getManagerData(loggeduser,None)
    else:
        redirect('/')
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Records'
    data = ('FNO', 'Member Number','Type of Menmbership','Company Name','Member Name','Contact Number','Email ID','Address','DOB','Membership Category','Date of Membership','Region','Status of Renewal','Date of Renewal','Renewal Amount')
    sheet.append(data)
    for record in allRecords:
        sheet.append((record.fno,record.member_no,record.membership_type,record.company_name,record.member_name,record.contact,record.email,record.address,record.dob,record.member_Category,record.membership_date,record.region,record.renewal_status,record.renewal_date,record.renewal_amount))
    filename = 'PixelStat_' + datetime.now().strftime("%Y_%m_%d - %H_%M_%S") + '.xlsx'
    wb.save(os.path.join(uploads_dir, secure_filename(filename)))
    path=os.path.join(uploads_dir, secure_filename(filename))
    return send_file(path,as_attachment=True)

# To remove alert message from web page
@app.route('/removeAlert',methods=['GET','POST'])
def removeAlert1():
    global alert
    alert = None
    return 'Done'

# Redirect to admin homepage
@app.route('/admin',methods=['GET','POST'])
def admin():
    global alert
    if 'loggeduser' in request.cookies:
        loggeduser = json.loads(request.cookies['loggeduser'])
        if loggeduser['user'] == 'admin':
            page = request.args.get('page',1,type=int)
            allRecords,allManagers,allregions,alert = getData(page)
            # if(len(allRecords) == 0):
            #     allRecords=[]
            return render_template('admin.html',allRecords=allRecords,allManagers=allManagers,allregions=allregions,current_date = datetime.now(),alert=alert)
    return redirect('/login')

# To add new record
@app.route('/admin/addRecord', methods=['GET','POST'])
@app.route('/manager/addRecord', methods=['GET','POST'])
def addRecord():
    global alert
    alert=None
    if request.method=='POST':
        mno = request.form['mno']
        cnm = request.form['cnm']
        mnm = request.form['mnm']
        contact = request.form['contact']
        email = request.form['email']
        address = request.form['address']
        dob = request.form['dob']
        category = request.form['category']
        membership_date = request.form['membership_date']
        
        try:
            dob = datetime.strptime(dob, '%Y-%m-%d')
        except:
            dob = None
        try:
            membership_date = datetime.strptime(membership_date, '%Y-%m-%d')
            # To find membership type
            days = datetime.now()-membership_date
            if days.days > (365*3):
                type='Individual A'
            else:
                type='Individual B'
        except:
            membership_date = None
            type = None
        
            
        region = request.form['region']
        renewal_status = request.form['renewal_status']
        renewal_status=False if renewal_status=='0' else True
        # fno = db.session.query(func.max(Excel_data.fno)).all()
        # print(fno)
        # if fno[0][0] == None:
        #     fno=1
        # else:
        #     fno=(fno[0][0])+1
        
        try:
            if int(renewal_status)==1:
                renewal_date = request.form['renewal_date']
                renewal_amount = request.form['renewal_amount']
                pixelstat=Excel_data(member_no=mno, membership_type=type,company_name=cnm,member_name=mnm,contact=contact,email=email,address=address,dob=dob,member_Category=category,membership_date=membership_date,region=region,renewal_status=renewal_status,renewal_date=datetime.strptime(renewal_date, '%Y-%m-%d'),renewal_amount=renewal_amount)
            else:
                pixelstat=Excel_data(member_no=mno, membership_type=type,company_name=cnm,member_name=mnm,contact=contact,email=email,address=address,dob=dob,member_Category=category,membership_date=membership_date,region=region,renewal_status=renewal_status)
                
            db.session.add(pixelstat)
            db.session.commit()
            if 'admin' in request.url_rule.rule:
                alert = {
                    'type':'success',
                    'message':'Record added successfully',
                    'section':'admin'
                }
            else:
                alert = {
                    'type':'success',
                    'message':'Record added successfully',
                    'section':'manager'
                }
                
        except exc.IntegrityError as err:
            print(err)
            if 'admin' in request.url_rule.rule:
                alert = {
                    'type':'danger',
                    'message':'Record insertion failed - Member No. should be unique',
                    'section':'admin'
                }
            else:
                alert = {
                    'type':'danger',
                    'message':'Record insertion failed - Member No. should be unique','section':'manager'
                }
        except exc.SQLAlchemyError as e:
            # print(e)
            if 'admin' in request.url_rule.rule:
                alert = {
                    'type':'danger',
                    'message':'Record insertion failed!!!',
                    'section':'admin'
                }
            else:
                alert = {
                    'type':'danger',
                    'message':'Record insertion failed!!!','section':'manager'
                }
        
        except:
            pass
            
    if 'admin' in request.url_rule.rule:
        return redirect('/admin#home')
    elif 'manager' in request.url_rule.rule:
        return redirect('/manager')

# To update record of excel
@app.route('/admin/updateRecord', methods=['GET','POST'])
@app.route('/manager/updateRecord', methods=['GET','POST'])
def updateRecord():
    global alert
    alert=None
    if request.method == 'POST':
        try:
            mno = request.form['mno']
            contact = request.form.get('contact')
            address = request.form.get('address')
            dob = request.form.get('dob')
            region = request.form.get('region')
            renewal_status = request.form['renewal_status']
            renewal_status=False if renewal_status=='0' else True       
            cnm = request.form.get('cnm')
            mnm = request.form.get('mnm')
            email = request.form.get('email')
            m_date = request.form.get('membership_date')
            category = request.form.get('category')
            type = None
            # To find membership type
            if(m_date):
                m_date = datetime.strptime(m_date, '%Y-%m-%d')
                days = datetime.now()-m_date
                if days.days > (365*3):
                    type='Individual A'
                else:
                    type='Individual B'
            if not mnm:
                mnm=0
            if not cnm:
                cnm=0
            if not dob:
                dob=None
            if not email:
                email=0
            if not contact:
                contact=0
            if not address:
                address=0
            if not m_date:
                m_date=None
            if not category:
                category=0
        except Exception as err:
            print(err)
        
        try:
            if dob:
                dob = datetime.strptime(dob, '%Y-%m-%d')
            
            if 'admin' in request.url_rule.rule:
                if int(renewal_status):
                    renewal_date = request.form.get('renewal_date')
                    renewal_amount = request.form.get('renewal_amount')
                    if not renewal_date:
                        renewal_date=None
                    if not renewal_amount:
                        renewal_amount=0
                    if renewal_date:
                        renewal_date = datetime.strptime(renewal_date, '%Y-%m-%d')
                    
                    Excel_data.query.filter_by(member_no=int(mno)).update(dict(membership_type=type,member_name=mnm,company_name=cnm,contact=contact,email=email,address=address,dob=dob,member_Category=category,membership_date=m_date,region=region,renewal_status=renewal_status,renewal_date=renewal_date,renewal_amount=renewal_amount))  
                else:
                    Excel_data.query.filter_by(member_no=int(mno)).update(dict(membership_type=type,member_name=mnm,company_name=cnm,contact=contact,email=email,address=address,dob=dob,member_Category=category,membership_date=m_date,region=region,renewal_status=renewal_status))  
                    
            elif 'manager' in request.url_rule.rule:
                if int(renewal_status):
                    renewal_date = request.form.get('renewal_date')
                    renewal_amount = request.form.get('renewal_amount')
                    if not renewal_date:
                        renewal_date=None
                    if not renewal_amount:
                        renewal_amount=0
                    if renewal_date:
                        renewal_date = datetime.strptime(renewal_date, '%Y-%m-%d')
                    Excel_data.query.filter_by(member_no=int(mno)).update(dict(contact=contact,company_name=cnm,address=address,dob=dob,region=region,renewal_status=renewal_status,renewal_date=renewal_date,renewal_amount=renewal_amount))  
                else:
                    Excel_data.query.filter_by(member_no=int(mno)).update(dict(company_name=cnm,contact=contact,address=address,dob=dob,region=region,renewal_status=renewal_status))  
            db.session.commit()
            if 'admin' in request.url_rule.rule:
                alert={
                    'type':'success',
                    'message':'Record updated successfully','section':'admin'
                }
            else:
                    alert={
                        'type':'success',
                        'message':'Record updated successfully','section':'manager'
                    }
                
        except exc.SQLAlchemyError as e:
            db.session.rollback()
            print(e)
            if 'admin' in request.url_rule.rule:
                alert={
                    'type':'danger',
                    'message':'Record updation failed','section':'admin'
                }
            else:
                alert={
                    'type':'danger',
                    'message':'Record updation failed','section':'manager'
                }
        except Exception as e:
            print(e)
            
    if 'admin' in request.url_rule.rule:
        return redirect('/admin#home')
    elif 'manager' in request.url_rule.rule:
        return redirect('/manager')    

# To delete record of excel
@app.route('/admin/deleteRecord', methods=['GET','POST'])
def deleteRecord():
    global alert
    alert=None
    if request.method=='POST':
        try:
            mno = request.form['id']
            pixelstat = Excel_data.query.filter_by(member_no=mno).first()
            db.session.delete(pixelstat)
            db.session.commit()
            alert={
                'type':'success',
                'message':'Record deleted successfully','section':'admin'
            }
        except:
            db.session.rollback()
            alert={
                'type':'danger',
                'message':'Something went wrong','section':'admin'
            }   
    return redirect('/admin#home')

# To get all data required for manager homepage
def getManagerData(loggeduser,page):
    manager = Managers.query.filter_by(mobile=int(loggeduser['username'])).all()
    regions = Region.query.filter_by(mid=manager[0].id).all()
    allregions = []
    for region in regions:
        allregions.append(region.region)
    if 'filter' in request.cookies:
        filter = json.loads(request.cookies['filter'])
        if (filter['region'] == None and filter['membership_type']==None and filter['member_no']==None):
            if page == None:
                allRecords = Excel_data.query.filter(Excel_data.region.in_(allregions)).all()
            else:
                allRecords = Excel_data.query.filter(Excel_data.region.in_(allregions)).paginate(page=page,per_page=4)
        else:
            filtered = {k: v for k, v in filter.items() if v is not None}
            filter.clear()
            filter.update(filtered)
            filteredData = db.session.query(Excel_data)
            for attr, value in filter.items():
                filteredData = filteredData.filter(getattr(Excel_data, attr).like("%%%s%%" % value),Excel_data.region.in_(allregions))
            if page == None:
                return filteredData,allregions,manager
            else:
                return filteredData.paginate(page=page,per_page=10),allregions,manager
    
    if page == None:
        allRecords = Excel_data.query.filter(Excel_data.region.in_(allregions)).all()
    else:
        allRecords = Excel_data.query.filter(Excel_data.region.in_(allregions)).paginate(page=page,per_page=10)
    return allRecords,allregions,manager

# Redirect to manager homepage
@app.route('/manager')
def manager():
    if 'loggeduser' in request.cookies:
        loggeduser = json.loads(request.cookies['loggeduser'])
        if loggeduser['user'] == 'manager':
            page = request.args.get('page',1,type=int)
            allRecords,allregions,manager = getManagerData(loggeduser,page)
            return render_template('manager.html',allRecords=allRecords,allregions=allregions,manager=manager[0],current_date = datetime.now(),alert=alert)
    return redirect('/login')
    
# To add a new manager
@app.route('/admin/addManager',methods=['GET','POST'])
def addManager():
    global alert
    alert=None
    if request.method == 'POST':
        name=request.form['name']
        mobile=request.form['mobile']
        email=request.form['email']
        password=request.form['password']
        regions=request.form.getlist('region')
        isdownloadable=request.form['downloadable']
        if isdownloadable=='No':
            isdownloadable=False
        else:
            isdownloadable = True

        try:
            pixelstat = Managers(name=name, mobile=mobile, email=email, password=password, able_to_download=isdownloadable)
            db.session.add(pixelstat)
            db.session.commit()
            for region in regions:
                regionobj = Region(region=region, mid=pixelstat.id)
                db.session.add(regionobj)
            db.session.commit()
            alert={
                'type':'success',
                'message':'Manager added successfully','section':'manager'
            }
            
        except exc.SQLAlchemyError as e:
            db.session.rollback()
            alert={
                'type':'danger',
                'message':'Something went wrong. Try again!','section':'manager'
            }
        except:
            pass
    return redirect('/admin#addManager')

# To edit manager record
@app.route('/admin/editManager',methods=['GET','POST'])
def updateManager():
    global alert
    alert=None
    if request.method == 'POST':
        id = request.form['mid']
        name=request.form['name']
        mobile=request.form['mobile']
        email=request.form['email']
        regions=request.form.getlist('region')
        isdownloadable=request.form['downloadable']
        if isdownloadable=='No':
            isdownloadable=False
        else:
            isdownloadable = True
        try:
            region = Region.query.filter_by(mid=id).delete()
            db.session.commit()
            Managers.query.filter_by(id=int(id)).update(dict(name=name,mobile=mobile,email=email,able_to_download=isdownloadable))
            db.session.commit()
            for region in regions:
                regionobj = Region(region=region, mid=id)
                db.session.add(regionobj)
            db.session.commit()
            alert={
                'type':'success',
                'message':'Manager updated successfully','section':'manager'
            }
        except exc.SQLAlchemyError as e:
            db.session.rollback()
            alert={
                'type':'danger',
                'message':'Something went wrong. Try again!','section':'manager'
            }
        except:
            pass
            
    return redirect('/admin#addManager')

# To delete manager record
@app.route('/admin/deleteManager/<int:id>')
def deleteManager(id):
    global alert
    alert=None
    try:
        manager = Managers.query.filter_by(id=id).first()
        db.session.delete(manager)
        db.session.commit()
        region = Region.query.filter_by(mid=id).delete()
        db.session.commit()
        alert={
                'type':'success',
                'message':'Manager deleted successfully','section':'manager'
            }
    except:
        db.session.rollback()
        alert={
                'type':'danger',
                'message':'Something went wrong. Try again!','section':'manager'
            }
        
    return redirect('/admin#addManager')

# To upload excel file store data in database
@app.route('/admin/uploadData',methods=['get','post'])
def uploadData():
    global alert
    alert=None
    if request.method == 'POST':
        excelfile = request.files['excelfile']
        excelfile.save(os.path.join(uploads_dir, secure_filename(excelfile.filename)))
        
        path = 'instance/uploads/'+excelfile.filename
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        row = sheet_obj.max_row
        # row = 100
        column = sheet_obj.max_column
        # column = 14
        inserted = False
        records = []
        
        count=0
        for i in range(2, row+1):
            record = []
            for j in range(2, column+1):
                cell_obj = sheet_obj.cell(row = i, column = j)
                record.append(cell_obj.value)
            # if len(record)==16 and record[1]!=None:
            if record[1]!=None:
                try:
                    r_date = datetime.strptime(record[12], '%d/%m/%Y')
                except:
                    r_date = None
                try:
                    dob = datetime.strptime(record[8], '%d/%m/%Y')
                except:
                    dob = None
                try:
                    m_date = datetime.strptime(record[10], '%d/%m/%Y')
                except:
                    m_date = None
                try:
                    # r_status = ('FALSE' if record[12]=='0' else record[12])
                    pixelstat=Excel_data(fno=record[0], member_no=record[1], membership_type=record[2],company_name=record[3],member_name=record[4],contact=record[5],email=record[6],address=record[7],dob=dob,member_Category=record[9],membership_date=m_date,region=record[11],renewal_date=r_date,renewal_amount=record[13])
                    db.session.add(pixelstat)
                    db.session.commit() 
                    inserted = True
                    count+=1
                except IntegrityError as e:
                    print(e)
                    db.session.rollback()
                except Exception as err:
                    print(err)
                    db.session.rollback()
                
        if inserted:
            alert={
                'type':'success',
                'message':'Data uploaded successfully','section':'admin'
            }
            return redirect('/admin#home')
        else:
            alert={
                'type':'danger',
                'message':'Something went wrong. Try again!','section':'upload-data'
            }
            return redirect('/admin#upload-data')                    
    return redirect('/admin')

# To fetch Login History of managers
@app.route('/getLoginHistory')
def getLoginHistory():
    managers = Managers.query.all()
    login_history = LoginHistory.query.all()
    
    managers_df = pd.DataFrame([(m.id, m.name, m.mobile, m.email) for m in managers], columns=['Manager ID', "Manager Name", "Manager Mobile", "Manager Email"])
    
    login_history_df = pd.DataFrame([(lh.mid, lh.login_time) for lh in login_history], columns=['Manager ID', 'Login Time'])
    
    merged_df = managers_df.merge(login_history_df, on="Manager ID", how='left')
    
    merged_df['Login Time'].fillna('N/A', inplace=True)
    
    excel_writer = pd.ExcelWriter('login_history.xlsx', engine='openpyxl')
    merged_df.to_excel(excel_writer, sheet_name='Managers_Login_History', index=False)
    
    excel_writer.close()
    
    return send_file('login_history.xlsx', as_attachment=True)
# # main function
# if __name__ == '__main__':
#     app.run(host="0.0.0.0",port=80)

# main function
if __name__ == '__main__':
    app.run(debug=True)